from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, BasePermission, AllowAny
from rest_framework.exceptions import ValidationError as DRFValidationError, PermissionDenied
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
import openpyxl
from datetime import timedelta

from .models import (
    Categoria, Instrumento, Usuario, Prestamo, 
    HistorialEstadoInstrumento, Perfil
)
from .serializers import (
    CategoriaSerializer,
    InstrumentoSerializer,
    InstrumentoDetailSerializer,
    UsuarioSerializer,
    PrestamoSerializer,
    HistorialEstadoInstrumentoSerializer,
    ReporteVencidoSerializer,
    ReporteEstadoInstrumentoSerializer,
    EstadisticasSerializer
)


# =========================
# PERMISOS PERSONALIZADOS
# =========================

class EsAdministrador(BasePermission):
    """Solo administrador tiene acceso."""
    def has_permission(self, request, view):
        return hasattr(request.user, 'perfil') and request.user.perfil.rol == 'administrador'


class EsAlmacenistaOAdministrador(BasePermission):
    """Almacenista o administrador tienen acceso."""
    def has_permission(self, request, view):
        return hasattr(request.user, 'perfil') and request.user.perfil.rol in ['administrador', 'almacenista']


class EsProfesorOSuperior(BasePermission):
    """Profesor, almacenista o administrador tienen acceso."""
    def has_permission(self, request, view):
        return hasattr(request.user, 'perfil') and request.user.perfil.rol in ['profesor', 'almacenista', 'administrador']


# =========================
# CATEGORIA
# =========================

class CategoriaViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar categorías de instrumentos."""
    
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [IsAuthenticated, EsAdministrador]

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'descripcion']
    ordering_fields = ['nombre']
    ordering = ['nombre']


# =========================
# INSTRUMENTO
# =========================

class InstrumentoViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar instrumentos."""
    
    queryset = Instrumento.objects.all()
    serializer_class = InstrumentoSerializer
    permission_classes = [IsAuthenticated]

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'referencia', 'marca', 'modelo', 'categoria__nombre']
    ordering_fields = ['nombre', 'fecha_adquisicion', 'estado']
    ordering = ['-fecha_adquisicion']

    def get_permissions(self):
        """
        GET (lectura): Cualquier usuario autenticado
        POST/PUT/DELETE (escritura): Solo almacenista o admin
        """
        if self.action in ['list', 'retrieve']:
            # Permitir lectura a cualquier autenticado
            permission_classes = [IsAuthenticated]
        elif self.action in ['historial']:
            # Historial de cambios: solo administrador
            permission_classes = [IsAuthenticated, EsAdministrador]
        else:
            # Restringir escritura a almacenista/admin
            permission_classes = [IsAuthenticated, EsAlmacenistaOAdministrador]
        
        return [permission() for permission in permission_classes]

    def get_serializer_class(self):
        """Usa serializer detallado para retrieve."""
        if self.action == 'retrieve':
            return InstrumentoDetailSerializer
        return InstrumentoSerializer

    def perform_create(self, serializer):
        """Crear instrumento: SOLO ADMINISTRADOR"""
        if not (hasattr(self.request.user, 'perfil') and self.request.user.perfil.rol == 'administrador'):
            raise PermissionDenied("No tienes permiso para crear instrumentos. Solo el administrador puede.")
        serializer.save()

    def perform_update(self, serializer):
        """Editar instrumento: ADMINISTRADOR y ALMACENISTA"""
        if not (hasattr(self.request.user, 'perfil') and self.request.user.perfil.rol in ['administrador', 'almacenista']):
            raise PermissionDenied("No tienes permiso para editar instrumentos.")
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Eliminar instrumento: SOLO ADMINISTRADOR"""
        if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'administrador'):
            return Response(
                {"error": "❌ No tienes permiso para eliminar instrumentos. Solo el administrador puede."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def dar_baja(self, request, pk=None):
        """Da de baja un instrumento (cambio de estado a 'baja')."""
        instrumento = self.get_object()

        if not (hasattr(request.user, 'perfil') and request.user.perfil.rol == 'administrador'):
            return Response(
                {"error": "No tienes permiso para dar de baja instrumentos. Solo administrador."},
                status=status.HTTP_403_FORBIDDEN
            )

        if instrumento.estado == 'prestado':
            return Response(
                {"error": "No se puede dar de baja un instrumento prestado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        instrumento.cambiar_estado(
            'baja',
            usuario=request.user,
            observacion=request.data.get('observacion', 'Instrumento dado de baja manualmente')
        )

        return Response({
            "mensaje": "Instrumento dado de baja correctamente.",
            "instrumento": InstrumentoDetailSerializer(instrumento).data
        })

    @action(detail=True, methods=['post'])
    def enviar_mantenimiento(self, request, pk=None):
        """Envía instrumento a mantenimiento."""
        instrumento = self.get_object()

        if instrumento.estado == 'prestado':
            return Response(
                {"error": "No se puede enviar a mantenimiento un instrumento prestado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        instrumento.cambiar_estado(
            'mantenimiento',
            usuario=request.user,
            observacion=request.data.get('observacion', 'Enviado a mantenimiento')
        )

        return Response({
            "mensaje": "Instrumento enviado a mantenimiento.",
            "instrumento": InstrumentoDetailSerializer(instrumento).data
        })

    @action(detail=True, methods=['post'])
    def cambiar_estado_fisico(self, request, pk=None):
        """
        Registra cambios en el estado físico de un instrumento después de devolución.
        Datos esperados: {'nuevo_estado': 'buen', 'observacion': '...'}
        Estados: disponible, mantenimiento, baja
        """
        instrumento = self.get_object()

        nuevo_estado = request.data.get('nuevo_estado')
        observacion = request.data.get('observacion', '')

        estados_validos = ['disponible', 'mantenimiento', 'baja']

        if nuevo_estado not in estados_validos:
            return Response(
                {"error": f"Estado inválido. Debe ser uno de: {', '.join(estados_validos)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        instrumento.cambiar_estado(
            nuevo_estado,
            usuario=request.user,
            observacion=f"Cambio de estado físico. {observacion}"
        )

        return Response({
            "mensaje": f"Estado del instrumento actualizado a: {nuevo_estado}",
            "instrumento": InstrumentoDetailSerializer(instrumento).data
        })

    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtiene el historial de movimientos de un instrumento."""
        instrumento = self.get_object()
        historial = instrumento.historial_movimientos.all().order_by('-fecha_cambio')
        serializer = HistorialEstadoInstrumentoSerializer(historial, many=True)
        return Response(serializer.data)


# =========================
# USUARIO
# =========================

class UsuarioViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar usuarios (estudiantes y profesores)."""
    
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated, EsAlmacenistaOAdministrador]

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'documento', 'correo', 'tipo']
    ordering_fields = ['nombre', 'tipo', 'fecha_creacion']
    ordering = ['nombre']


# =========================
# PRESTAMO
# =========================

class PrestamoViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar préstamos de instrumentos."""
    
    queryset = Prestamo.objects.all()
    serializer_class = PrestamoSerializer
    permission_classes = [IsAuthenticated]

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['instrumento__nombre', 'usuario__nombre', 'usuario__documento', 'estado']
    ordering_fields = ['fecha_prestamo', 'estado', 'fecha_vencimiento']
    ordering = ['-fecha_prestamo']

    def get_permissions(self):
        """
        GET (lectura): Cualquier usuario autenticado
        POST/PUT/DELETE (escritura): Solo almacenista o admin
        """
        if self.action in ['list', 'retrieve', 'vencidos']:
            # Permitir lectura a cualquier autenticado
            permission_classes = [IsAuthenticated]
        else:
            # Restringir escritura a almacenista/admin
            permission_classes = [IsAuthenticated, EsAlmacenistaOAdministrador]
        
        return [permission() for permission in permission_classes]

    @action(detail=True, methods=['post'])
    def devolver(self, request, pk=None):
        """
        Cierra un préstamo y devuelve el instrumento.
        Permisos: ADMINISTRADOR y ALMACENISTA
        Datos opcionales: {'observacion': '...'}
        """
        # Verificar permisos
        if not (hasattr(request.user, 'perfil') and request.user.perfil.rol in ['almacenista', 'administrador']):
            return Response(
                {"error": "❌ No tienes permiso para devolver instrumentos. Solo almacenista y administrador."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        prestamo = self.get_object()

        if prestamo.estado == 'disponible':
            return Response(
                {"error": "Este préstamo ya está cerrado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Actualizar préstamo
        prestamo.estado = 'disponible'
        prestamo.fecha_devolucion = timezone.now().date()
        prestamo._usuario_sistema = request.user
        prestamo.save()

        # Cambiar estado del instrumento a disponible
        prestamo.instrumento.cambiar_estado(
            'disponible',
            usuario=request.user,
            observacion=request.data.get('observacion', 'Instrumento devuelto')
        )

        return Response({
            "mensaje": "✅ Préstamo cerrado correctamente.",
            "prestamo": PrestamoSerializer(prestamo).data
        })

    @action(detail=False, methods=['get'])
    def vencidos(self, request):
        """
        Obtiene la lista de préstamos vencidos que no han sido devueltos.
        Alerta a almacenista sobre instrumentos no devueltos en tiempo.
        """
        prestamos_vencidos = Prestamo.prestamos_vencidos()
        serializer = ReporteVencidoSerializer(prestamos_vencidos, many=True)
        
        return Response({
            'cantidad': len(prestamos_vencidos),
            'prestamos': serializer.data
        })

    @action(detail=False, methods=['get'])
    def proximos_a_vencer(self, request):
        """
        Obtiene préstamos que vencen en los próximos 7 días.
        """
        hoy = timezone.now().date()
        fecha_limite = hoy + timedelta(days=7)

        prestamos = Prestamo.objects.filter(
            estado='enuso',
            fecha_vencimiento__gte=hoy,
            fecha_vencimiento__lte=fecha_limite,
            fecha_devolucion__isnull=True
        ).order_by('fecha_vencimiento')

        serializer = ReporteVencidoSerializer(prestamos, many=True)
        
        return Response({
            'cantidad': len(prestamos),
            'prestamos': serializer.data
        })

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exporta todos los préstamos a un archivo Excel."""
        prestamos = Prestamo.objects.all().select_related('instrumento', 'usuario')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prestamos"

        # Encabezados
        ws.append([
            "ID",
            "Instrumento",
            "Referencia",
            "Usuario",
            "Documento",
            "Fecha Préstamo",
            "Fecha Vencimiento",
            "Fecha Devolución",
            "Días Permitidos",
            "Estado",
            "Observaciones"
        ])

        # Datos
        for p in prestamos:
            ws.append([
                p.id,
                p.instrumento.nombre,
                p.instrumento.referencia,
                p.usuario.nombre,
                p.usuario.documento,
                str(p.fecha_prestamo) if p.fecha_prestamo else "",
                str(p.fecha_vencimiento) if p.fecha_vencimiento else "",
                str(p.fecha_devolucion) if p.fecha_devolucion else "",
                p.dias_permitidos,
                p.estado,
                p.observaciones or ""
            ])

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 20

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=prestamos_reporte.xlsx'

        wb.save(response)

        return response

    @action(detail=False, methods=['get'])
    def reporte_estado_instrumentos(self, request):
        """
        Genera reporte detallado sobre el estado de los instrumentos.
        """
        instrumentos = Instrumento.objects.all()
        serializer = ReporteEstadoInstrumentoSerializer(instrumentos, many=True)

        # Estadísticas por estado
        stats_por_estado = {
            'disponible': instrumentos.filter(estado='disponible').count(),
            'prestado': instrumentos.filter(estado='prestado').count(),
            'mantenimiento': instrumentos.filter(estado='mantenimiento').count(),
            'baja': instrumentos.filter(estado='baja').count(),
        }

        return Response({
            'total_instrumentos': instrumentos.count(),
            'estadisticas_por_estado': stats_por_estado,
            'instrumentos': serializer.data
        })


# =========================
# REPORTES Y ESTADÍSTICAS
# =========================

class ReportesViewSet(viewsets.ViewSet):
    """ViewSet para reportes y estadísticas del sistema."""
    
    permission_classes = [IsAuthenticated, EsAlmacenistaOAdministrador]

    @action(detail=False, methods=['get'])
    def estadisticas_generales(self, request):
        """Obtiene estadísticas generales del sistema."""
        total_instrumentos = Instrumento.objects.count()
        instrumentos_disponibles = Instrumento.objects.filter(estado='disponible').count()
        instrumentos_prestados = Instrumento.objects.filter(estado='prestado').count()
        instrumentos_mantenimiento = Instrumento.objects.filter(estado='mantenimiento').count()
        instrumentos_baja = Instrumento.objects.filter(estado='baja').count()

        total_prestamos_activos = Prestamo.objects.filter(estado='enuso').count()
        prestamos_vencidos = Prestamo.prestamos_vencidos().count()

        total_usuarios = Usuario.objects.count()
        total_categorias = Categoria.objects.count()

        datos = {
            'total_instrumentos': total_instrumentos,
            'instrumentos_disponibles': instrumentos_disponibles,
            'instrumentos_prestados': instrumentos_prestados,
            'instrumentos_mantenimiento': instrumentos_mantenimiento,
            'instrumentos_baja': instrumentos_baja,
            'total_prestamos_activos': total_prestamos_activos,
            'prestamos_vencidos': prestamos_vencidos,
            'total_usuarios': total_usuarios,
            'total_categorias': total_categorias
        }

        serializer = EstadisticasSerializer(datos)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def reporte_uso_instrumentos(self, request):
        """Reporte de uso de instrumentos (cuántas veces se prestan, etc)."""
        instrumentos = Instrumento.objects.annotate(
            total_prestamos=Count('prestamo', filter=Q(prestamo__estado='disponible')),
            prestamos_activos=Count('prestamo', filter=Q(prestamo__estado='enuso'))
        )

        data = []
        for instr in instrumentos:
            data.append({
                'id': instr.id,
                'nombre': instr.nombre,
                'referencia': instr.referencia,
                'categoria': instr.categoria.nombre,
                'estado': instr.estado,
                'total_prestamos': instr.total_prestamos,
                'prestamos_activos': instr.prestamos_activos,
            })

        return Response({
            'total_instrumentos': len(data),
            'reportes': data
        })

    @action(detail=False, methods=['get'])
    def reporte_usuarios_morosos(self, request):
        """Reporte de usuarios con préstamos vencidos."""
        prestamos_vencidos = Prestamo.prestamos_vencidos()
        
        usuarios_morosos = {}
        for prestamo in prestamos_vencidos:
            usuario_id = prestamo.usuario.id
            if usuario_id not in usuarios_morosos:
                usuarios_morosos[usuario_id] = {
                    'usuario': {
                        'id': prestamo.usuario.id,
                        'nombre': prestamo.usuario.nombre,
                        'documento': prestamo.usuario.documento,
                        'correo': prestamo.usuario.correo,
                        'telefono': prestamo.usuario.telefono,
                    },
                    'prestamos_vencidos': []
                }
            
            usuarios_morosos[usuario_id]['prestamos_vencidos'].append({
                'instrumento': prestamo.instrumento.nombre,
                'referencia': prestamo.instrumento.referencia,
                'fecha_prestamo': str(prestamo.fecha_prestamo),
                'fecha_vencimiento': str(prestamo.fecha_vencimiento),
                'dias_vencido': (timezone.now().date() - prestamo.fecha_vencimiento).days
            })

        return Response({
            'total_usuarios_morosos': len(usuarios_morosos),
            'usuarios': list(usuarios_morosos.values())
        })

    @action(detail=False, methods=['get'])
    def reporte_detalles_prestamos(self, request):
        """Reporte detallado de todos los préstamos con información completa."""
        prestamos = Prestamo.objects.all().select_related('instrumento', 'usuario')

        data = []
        for p in prestamos:
            dias_vencimiento = None
            estado_vencimiento = 'vigente'

            if p.fecha_vencimiento:
                dias_vencimiento = (timezone.now().date() - p.fecha_vencimiento).days
                if dias_vencimiento > 0:
                    estado_vencimiento = 'vencido'
                elif dias_vencimiento >= -7:
                    estado_vencimiento = 'proximo_vencer'

            data.append({
                'id': p.id,
                'instrumento': p.instrumento.nombre,
                'referencia': p.instrumento.referencia,
                'usuario': p.usuario.nombre,
                'documento': p.usuario.documento,
                'fecha_prestamo': str(p.fecha_prestamo),
                'fecha_vencimiento': str(p.fecha_vencimiento),
                'fecha_devolucion': str(p.fecha_devolucion) if p.fecha_devolucion else None,
                'estado': p.estado,
                'dias_permitidos': p.dias_permitidos,
                'dias_transcurridos': (timezone.now().date() - p.fecha_prestamo).days if p.fecha_prestamo else 0,
                'dias_vencimiento': dias_vencimiento,
                'estado_vencimiento': estado_vencimiento,
                'observaciones': p.observaciones
            })

        return Response({
            'total_prestamos': len(data),
            'prestamos': data
        })

    @action(detail=False, methods=['get'])
    def exportar_reporte_excel(self, request):
        """Exporta reporte completo a Excel con múltiples hojas."""
        # Crear workbook con múltiples hojas
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Hoja 1: Estadísticas Generales
        ws1 = wb.create_sheet("Estadísticas")
        ws1.append(["Concepto", "Cantidad"])
        
        total_instrumentos = Instrumento.objects.count()
        disponibles = Instrumento.objects.filter(estado='disponible').count()
        prestados = Instrumento.objects.filter(estado='prestado').count()
        mantenimiento = Instrumento.objects.filter(estado='mantenimiento').count()
        baja = Instrumento.objects.filter(estado='baja').count()
        total_prestamos_activos = Prestamo.objects.filter(estado='enuso').count()
        vencidos = Prestamo.prestamos_vencidos().count()

        ws1.append(["Total de Instrumentos", total_instrumentos])
        ws1.append(["Instrumentos Disponibles", disponibles])
        ws1.append(["Instrumentos Prestados", prestados])
        ws1.append(["Instrumentos en Mantenimiento", mantenimiento])
        ws1.append(["Instrumentos dados de Baja", baja])
        ws1.append(["Total Préstamos Activos", total_prestamos_activos])
        ws1.append(["Préstamos Vencidos", vencidos])

        # Hoja 2: Préstamos
        ws2 = wb.create_sheet("Prestamos")
        ws2.append([
            "Instrumento", "Referencia", "Usuario", "Doc.", 
            "Fecha Préstamo", "Vencimiento", "Devolución", "Estado"
        ])

        for p in Prestamo.objects.all().select_related('instrumento', 'usuario'):
            ws2.append([
                p.instrumento.nombre,
                p.instrumento.referencia,
                p.usuario.nombre,
                p.usuario.documento,
                str(p.fecha_prestamo) if p.fecha_prestamo else "",
                str(p.fecha_vencimiento) if p.fecha_vencimiento else "",
                str(p.fecha_devolucion) if p.fecha_devolucion else "",
                p.estado
            ])

        # Hoja 3: Instrumentos
        ws3 = wb.create_sheet("Instrumentos")
        ws3.append([
            "Nombre", "Referencia", "Categoría", "Estado", 
            "Marca", "Modelo", "Ubicación", "Fecha Adquisición"
        ])

        for i in Instrumento.objects.all().select_related('categoria'):
            ws3.append([
                i.nombre,
                i.referencia,
                i.categoria.nombre,
                i.estado,
                i.marca or "",
                i.modelo or "",
                i.ubicacion_fisica or "",
                str(i.fecha_adquisicion) if i.fecha_adquisicion else ""
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=reporte_completo.xlsx'

        wb.save(response)
        return response


# =========================
# ENDPOINTS DE AUTENTICACIÓN
# =========================

@ensure_csrf_cookie  # ✅ Asegurar que Django coloque la cookie del CSRF token
@api_view(['GET'])
@permission_classes([AllowAny])
def obtener_csrf_token(request):
    """
    Retorna el CSRF token para solicitudes POST/PUT/DELETE.
    Este endpoint debe llamarse antes del login.
    """
    from django.middleware.csrf import get_token
    csrf_token = get_token(request)
    return Response({
        'csrfToken': csrf_token
    }, status=status.HTTP_200_OK)


@csrf_protect  # ✅ Validar CSRF en POST
@api_view(['POST'])
@permission_classes([AllowAny])
def login_personalizado(request):
    """
    Endpoint de login con SESIONES (en lugar de tokens).
    Espera: { "username": "...", "password": "..." }
    Retorna: { "username": "...", "rol": "...", "user_id": ... }
    
    La sesión se mantiene automáticamente en cookies del navegador.
    👮 ACCESO RESTRINGIDO: Solo administrador y almacenista pueden entrar
    """
    
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response({
            'error': 'Se requieren username y password'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Autenticar usuario
    user = authenticate(username=username, password=password)
    
    if not user:
        return Response({
            'error': 'Credenciales inválidas'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    # Verificar que el usuario sea administrador o almacenista
    rol = 'usuario'
    tiene_acceso = False

    # Permitir acceso inmediato a superusuarios/staff aunque no tengan perfil creado.
    if user.is_superuser or user.is_staff:
        rol = 'administrador'
        tiene_acceso = True
    
    try:
        if hasattr(user, 'perfil'):
            rol = user.perfil.rol
            # Solo estos roles pueden acceder
            if rol in ['administrador', 'almacenista']:
                tiene_acceso = True
    except Perfil.DoesNotExist:
        rol = 'usuario'
    
    # Si no es admin o almacenista, rechazar
    if not tiene_acceso:
        return Response({
            'error': f'Acceso denegado. Solo administrador y almacenista pueden entrar. Tu rol es: {rol}'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # ✅ Usar login de Django (sesión)
    login(request, user)
    
    return Response({
        'user_id': user.id,
        'username': user.username,
        'email': user.email,
        'rol': rol,
        'mensaje': f'Login exitoso como {rol}'
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_usuario_actual(request):
    """
    Retorna la información del usuario actual autenticado en sesión.
    Valida contra la base de datos cada vez que se llama.
    """
    user = request.user
    
    if not user.is_authenticated:
        return Response({
            'error': 'No autenticado'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    rol = 'usuario'
    try:
        if hasattr(user, 'perfil'):
            rol = user.perfil.rol
    except Perfil.DoesNotExist:
        rol = 'usuario'
    
    return Response({
        'user_id': user.id,
        'username': user.username,
        'email': user.email,
        'rol': rol,
        'is_authenticated': True
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def logout_personalizado(request):
    """
    Cierra la sesión del usuario actual.
    Es idempotente: si no hay sesión activa, responde 200 igualmente.
    """
    if request.user.is_authenticated:
        logout(request)
        mensaje = 'Sesión cerrada correctamente'
    else:
        mensaje = 'No había una sesión activa'

    return Response({
        'mensaje': mensaje
    }, status=status.HTTP_200_OK)